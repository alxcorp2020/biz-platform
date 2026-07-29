#!/usr/bin/env python3
"""1차 버전: 첨부파일 텍스트 추출 배치 스크립트.

attachments.extraction_status='pending'인 행을 조회해, 파일 형식별로
텍스트만 뽑아 extracted_text/extraction_status에 기록한다. 큐나 스케줄러는
없다 — 실행하면 그 시점의 pending 건을 전부(또는 --limit만큼) 처리하고
끝난다. Go 쪽과의 실시간 연동, 그리고 자격조건/제출서류 등 구조화 추출은
다음 단계.

사용법:
    venv/bin/python run_extraction.py [--limit N] [--attachment-dir DIR]

환경변수:
    DATABASE_URL     Postgres 연결 문자열 (collector와 동일한 값 사용 가능)
    ATTACHMENT_DIR   첨부파일이 저장된 로컬 디렉터리 (기본 ./data/attachments —
                     collector/internal/collector/runner의 기본값과 일치시킬 것)
"""
import argparse
import logging
import os
import subprocess
import sys

import psycopg2

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("run_extraction")

# 지원 형식. 나머지(zip, egg, docx, pptx, dwg 등)는 1차 버전에서 미지원으로
# 표시만 하고 넘어간다 — 특히 .egg는 국내 전용 암호화 압축 포맷이라 공개된
# 파이썬 라이브러리가 없다.
SUPPORTED_TYPES = {"hwp", "hwpx", "pdf", "xlsx"}


class ExtractionError(Exception):
    pass


def extract_hwp(path):
    """pyhwp는 안정적인 공개 파이썬 API가 없어(내부 모델이 복잡하고 문서화가
    부실함), 이미 동작이 검증된 hwp5txt CLI를 서브프로세스로 호출한다."""
    hwp5txt = os.path.join(os.path.dirname(sys.executable), "hwp5txt")
    out_path = path + ".extracted.txt"
    try:
        proc = subprocess.run(
            [hwp5txt, path, "--output", out_path],
            capture_output=True, text=True, timeout=60,
        )
        if proc.returncode != 0:
            raise ExtractionError(f"hwp5txt exit {proc.returncode}: {proc.stderr.strip()[:500]}")
        with open(out_path, "r", encoding="utf-8") as f:
            return f.read()
    finally:
        if os.path.exists(out_path):
            os.remove(out_path)


def extract_hwpx(path):
    from hwp_hwpx_parser import extract_hwpx as _extract_hwpx
    result = _extract_hwpx(path)
    text = result[0] if isinstance(result, tuple) else getattr(result, "text", None)
    if not text:
        raise ExtractionError("hwp_hwpx_parser returned no text")
    return text


def _render_pdf_table(table):
    """표를 '<표>' 마커 + '셀1 | 셀2 | 셀3' 평문으로 렌더링한다. HWP/HWPX
    파서가 이미 남기는 '<표>' 마커와 형식을 맞춰, 다음 단계(섹션 추출)가
    포맷 무관하게 "표 근처인지"를 판단할 수 있게 한다."""
    lines = ["<표>"]
    for row in table:
        cells = [(c or "").strip().replace("\n", " ") for c in row]
        lines.append(" | ".join(cells))
    return "\n".join(lines)


def extract_pdf(path):
    import pdfplumber
    parts = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
            for table in page.extract_tables():
                parts.append(_render_pdf_table(table))
    if not parts:
        raise ExtractionError("no extractable text (스캔 이미지 PDF일 가능성)")
    return "\n\n".join(parts)


def extract_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            line = " ".join(str(c) for c in row if c is not None).strip()
            if line:
                parts.append(line)
    if not parts:
        raise ExtractionError("no non-empty cells found")
    return "\n".join(parts)


EXTRACTORS = {
    "hwp": extract_hwp,
    "hwpx": extract_hwpx,
    "pdf": extract_pdf,
    "xlsx": extract_xlsx,
}


def process_one(conn, attachment_dir, att_id, stored_filename, file_type):
    file_type = (file_type or "").lower()
    path = os.path.join(attachment_dir, stored_filename)

    if file_type not in SUPPORTED_TYPES:
        return "unsupported", None, f"미지원 형식: {file_type or '(확장자 없음)'}"

    if not os.path.exists(path):
        return "failed", None, f"파일이 디스크에 없음: {path}"

    try:
        text = EXTRACTORS[file_type](path)
        # Postgres TEXT는 NUL(0x00) 바이트를 저장할 수 없다 — 구형 hwp
        # 바이너리 파싱 결과에 간혹 섞여 나온다.
        text = text.replace("\x00", "")
        return "completed", text, None
    except Exception as e:  # noqa: broad-except — 어떤 라이브러리든 실패 사유를 그대로 기록
        return "failed", None, f"{type(e).__name__}: {e}"


def main():
    parser = argparse.ArgumentParser(description="첨부파일 텍스트 추출 배치")
    parser.add_argument("--limit", type=int, default=None, help="처리할 최대 건수 (기본: 전체)")
    parser.add_argument("--attachment-dir", default=os.environ.get("ATTACHMENT_DIR", "./data/attachments"))
    args = parser.parse_args()

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        logger.error("DATABASE_URL 환경변수가 설정되어 있지 않습니다")
        sys.exit(1)

    conn = psycopg2.connect(dsn)
    conn.autocommit = False

    query = "SELECT id, stored_filename, file_type FROM attachments WHERE extraction_status = 'pending' ORDER BY created_at"
    if args.limit:
        query += f" LIMIT {int(args.limit)}"

    with conn.cursor() as cur:
        cur.execute(query)
        rows = cur.fetchall()

    logger.info("처리 대상: %d건 (attachment_dir=%s)", len(rows), args.attachment_dir)

    counts = {"completed": 0, "failed": 0, "unsupported": 0}
    for att_id, stored_filename, file_type in rows:
        status, text, error = process_one(conn, args.attachment_dir, att_id, stored_filename, file_type)

        try:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE attachments SET extraction_status = %s, extracted_text = %s, extraction_error = %s WHERE id = %s",
                    (status, text, error, att_id),
                )
            conn.commit()
        except Exception as e:  # noqa: broad-except — 한 건의 저장 실패로 전체 배치가 죽으면 안 됨
            conn.rollback()
            status, error = "failed", f"DB 저장 실패: {type(e).__name__}: {e}"
            logger.error("attachment %s 저장 실패: %s", att_id, error)

        counts[status] += 1
        if status == "completed":
            logger.info("완료 [%s] %s (%d자)", file_type, stored_filename, len(text))
        else:
            logger.warning("%s [%s] %s — %s", status, file_type, stored_filename, error)

    conn.close()
    logger.info("종료: 완료 %d / 실패 %d / 미지원 %d",
                counts["completed"], counts["failed"], counts["unsupported"])


if __name__ == "__main__":
    main()
